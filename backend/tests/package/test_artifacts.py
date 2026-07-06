"""Pure artifact builder tests — ``letter.docx`` + ``chronology.xlsx`` (M5 Wave B2).

Self-contained (own in-memory engine + firm/attorney/matter, direct ORM), matching
``tests/package/test_manifest.py``'s fixture style. Synthetic data only — no PHI. DraftSection rows
are built BY HAND (rendered_preview + spans hand-made), never via the drafter (the sibling Brain-2
wave is not imported here).

Coverage: docx round-trips (headings + paragraph text via python-docx); the memo is excluded from
the letter; a planted token in a rendered_preview raises ArtifactTokenLeak; docx + xlsx byte
determinism (build twice -> identical sha256); xlsx header + cell values exact; xlsx narrative-cell
token scan.
"""

from __future__ import annotations

import hashlib
import io
import uuid

import pytest
from openpyxl import load_workbook

from app.models.enums import SectionValidation
from app.models.orm import DraftSection
from app.package import artifacts


def _section(
    *, section_id: str, rendered_preview: str | None, sort_order: int = 0, spans: list | None = None
) -> DraftSection:
    """A hand-built PASSED DraftSection row (no drafter, no DB) for the pure builders."""
    return DraftSection(
        id=uuid.uuid4(),
        firm_id=uuid.uuid4(),
        draft_id=uuid.uuid4(),
        section_id=section_id,
        purpose="p",
        body_tokenized="x",
        rendered_preview=rendered_preview,
        registry_version=1,
        validation=SectionValidation.PASSED.value,
        spans=spans or [],
        sort_order=sort_order,
    )


# --------------------------------------------------------------------------------------
# build_letter_docx
# --------------------------------------------------------------------------------------


def _docx_paragraph_texts(data: bytes) -> list[str]:
    from docx import Document

    document = Document(io.BytesIO(data))
    return [p.text for p in document.paragraphs]


def test_letter_docx_has_letterhead_re_line_and_section_content() -> None:
    sections = [
        _section(
            section_id="liability_summary",
            rendered_preview="The defendant ran the light.\n\nDamages follow.",
            sort_order=1,
        ),
    ]
    data = artifacts.build_letter_docx(
        firm_name="Acme Law", client_display_name="Jane Doe", sections=sections
    )
    texts = _docx_paragraph_texts(data)

    assert "Acme Law" in texts  # generated letterhead heading
    assert "Re: Jane Doe" in texts
    # section heading derived from section_id (title-cased, underscores -> spaces)
    assert "Liability Summary" in texts
    # rendered_preview split on blank lines into paragraphs
    assert "The defendant ran the light." in texts
    assert "Damages follow." in texts


def test_letter_docx_excludes_memo() -> None:
    sections = [_section(section_id="intro", rendered_preview="Body.", sort_order=1)]
    data = artifacts.build_letter_docx(
        firm_name="Acme Law",
        client_display_name="Jane Doe",
        sections=sections,
        memo="SECRET STRATEGY MEMO — never on the wire",
    )
    joined = "\n".join(_docx_paragraph_texts(data))
    assert "SECRET STRATEGY MEMO" not in joined


def test_letter_docx_token_leak_raises() -> None:
    # A token planted in the rendered_preview (a resolution bug) must fail the build.
    sections = [
        _section(
            section_id="damages", rendered_preview="Total billed was [[FACT_1]].", sort_order=1
        )
    ]
    with pytest.raises(artifacts.ArtifactTokenLeak) as excinfo:
        artifacts.build_letter_docx(
            firm_name="Acme Law", client_display_name="Jane Doe", sections=sections
        )
    assert excinfo.value.token == "[[FACT_1]]"
    assert excinfo.value.section_id == "damages"


def test_letter_docx_is_byte_deterministic() -> None:
    sections = [
        _section(section_id="intro", rendered_preview="Para one.\n\nPara two.", sort_order=1),
        _section(section_id="close", rendered_preview="Closing.", sort_order=2),
    ]
    kwargs = dict(firm_name="Acme Law", client_display_name="Jane Doe", sections=sections)
    a = artifacts.build_letter_docx(**kwargs)
    b = artifacts.build_letter_docx(**kwargs)
    assert hashlib.sha256(a).hexdigest() == hashlib.sha256(b).hexdigest()


def test_letter_docx_orders_sections_as_given() -> None:
    # The builder emits sections in the order it is handed (the caller sorts by sort_order).
    sections = [
        _section(section_id="first_section", rendered_preview="One.", sort_order=1),
        _section(section_id="second_section", rendered_preview="Two.", sort_order=2),
    ]
    texts = _docx_paragraph_texts(
        artifacts.build_letter_docx(
            firm_name="Acme Law", client_display_name="Jane Doe", sections=sections
        )
    )
    assert texts.index("First Section") < texts.index("Second Section")


# --------------------------------------------------------------------------------------
# build_chronology_xlsx
# --------------------------------------------------------------------------------------


def _rows() -> list[dict]:
    return [
        {
            "row_id": str(uuid.uuid4()),
            "date_of_service": "2026-01-10",
            "provider_display": "Dr. Alice",
            "facility_display": "General Hospital",
            "encounter_type": "ER",
            "narrative": "Patient presented with neck pain.",
        },
        {
            "row_id": str(uuid.uuid4()),
            "date_of_service": "2026-02-01",
            "provider_display": "Dr. Bob",
            "facility_display": "PT Clinic",
            "encounter_type": "PT",
            "narrative": "Physical therapy, first session.",
        },
    ]


def test_chronology_xlsx_header_and_cells_exact() -> None:
    data = artifacts.build_chronology_xlsx(_rows())
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["Date of service", "Provider", "Facility", "Type", "Narrative"]
    # First data row.
    assert [c.value for c in ws[2]] == [
        "2026-01-10",
        "Dr. Alice",
        "General Hospital",
        "ER",
        "Patient presented with neck pain.",
    ]
    # Second data row narrative.
    assert ws.cell(row=3, column=5).value == "Physical therapy, first session."


def test_chronology_xlsx_token_leak_raises() -> None:
    rows = _rows()
    rows[1]["narrative"] = "See [[AMT_2]] for the bill."
    with pytest.raises(artifacts.ArtifactTokenLeak) as excinfo:
        artifacts.build_chronology_xlsx(rows)
    assert excinfo.value.token == "[[AMT_2]]"


def test_chronology_xlsx_is_byte_deterministic() -> None:
    a = artifacts.build_chronology_xlsx(_rows())
    b = artifacts.build_chronology_xlsx(_rows())
    assert hashlib.sha256(a).hexdigest() == hashlib.sha256(b).hexdigest()


def test_chronology_xlsx_empty_rows_builds_header_only() -> None:
    data = artifacts.build_chronology_xlsx([])
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.max_row == 1  # header only
    assert [c.value for c in ws[1]] == [
        "Date of service",
        "Provider",
        "Facility",
        "Type",
        "Narrative",
    ]


def test_chronology_xlsx_stringifies_missing_keys() -> None:
    # A row missing keys renders empty cells, never a KeyError. (openpyxl reads a written empty
    # string back as ``None`` — the point is that the row builds and the narrative lands.)
    data = artifacts.build_chronology_xlsx([{"narrative": "only narrative"}])
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert [c.value for c in ws[2]] == [None, None, None, None, "only narrative"]
